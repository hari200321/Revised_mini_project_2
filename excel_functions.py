"""
excel_functions.py
"""

from openpyxl import load_workbook

class Hari_Excel_Functions:

    def __init__(self, excel_file_name, sheet_name):

        self.file = excel_file_name
        self.sheet = sheet_name

    #fetch the total row count
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    #fetch the total column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    #read data from particular row and column
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    #write the data into particular row and column
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value=data
        workbook.save(self.file)
